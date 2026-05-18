"""
Petice Analyzer — ruceprycodmedii.cz
======================================
Načte CSV z petice_scraper.py a provede analýzu:
  - Sprostá slova
  - Podezřelá slova
  - Slovesa v jméně/městě/povolání
  - Podezřelá délka
  - Duplicity (jméno + město + povolání)

Závislosti:
    pip install openpyxl

Použití:
    python petice_analyzer.py
"""

import csv
import re
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# ============================================================================
# KONFIG — prahy podezřelé délky
# ============================================================================

MIN_JMENO = 6    # jméno+příjmení kratší než toto → podezřelé
MAX_JMENO = 35   # jméno+příjmení delší než toto → podezřelé
MAX_MESTO = 50   # město delší než toto → podezřelé

RUN_ID      = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_XLSX = f"petice_problemy_{RUN_ID}.xlsx"

# ============================================================================
# NORMALIZACE
# ============================================================================

def strip_diacritics(text: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    )

def normalize(text: str) -> str:
    if not text:
        return ""
    text = str(text).lower().strip()
    text = strip_diacritics(text)
    text = re.sub(r"\s+", " ", text)
    return text

# ============================================================================
# PROFANITY LIST — substring matching (chytí i uprostřed slova)
# ============================================================================

PROFANITY_LIST = [
    # České — dlouhé
    "hovno", "hovado", "hovnivous",
    "kurva", "kurvy", "kurvo",
    "pička", "picka", "pičus", "picuska", "píča", "pýča", "piče",
    "kokot", "kokotina",
    "zmrd", "mrdka", "mrdas", "mrdac", "mrdalpsa", "mrdat",
    "sráč", "srač", "nasra", "vysra", "posra", "zasra", 
    "hajzl", "hajzle",
    "čurák", "curak", "čůrák", "čůrak",
    "kreténe", "kreten",
    "šlapka", "slapka",
    "nouma", "ňouma",
    "zmetek",
    "bastard",
    "debil",
    "idiot",
    "posera", "podela",
    "hovn",
    "hovad",
    "penis",
    "anal",
    "pizd",
    "stet",
    "parchant",
    "ichtyl",
    "demen", "imbeci", "deges",
    "pich",
    "soust", "ojet", "ohnou", "ohni",
    "kour", "plyn", "deport",
    # Krátké české — dostanou hranice automaticky
    "vůl", "vole",
    "prd", "prdel",
    # Anglické
    "fuck", "dick", "suck", "shit", "bullshit", "asshole", "bitch", "cunt", "wanker", "drug", "deal"
]

SHORT_PROFANITY = {"ass", "vul", "vole", "prd", "prdel", "pic", "rit", "suk", "drz", "hul", "kur", "zid"}
# ============================================================================
# WHITELIST — výjimky z profanity (přidávej sem další)
# ============================================================================

# Pokud najdeme toto slovo A město odpovídá → ignorujeme
PROFANITY_CITY_WHITELIST = {
    "pice": {"krepice", "ropice", "repice", "upice", "jistebnice", "holice", "cepice", "pocepice", "jarpice", "popice", "herspice", "biskupice"},
    "stet": {"steti", "stetovice", "hostetin"},
    "piče": {"krepice", "ropice", "repice", "upice", "jistebnice", "holice", "cepice", "pocepice", "jarpice", "popice", "herspice", "biskupice"},
    "ojet": {"kojetin", "svojetice", "kojetice"},
    # "dalsi_slovo": {"město1", "město2"},
}

# Pokud najdeme toto slovo A příjmení obsahuje povolený vzor → ignorujeme
PROFANITY_NAME_WHITELIST = {
    "pička": {"špička", "krupička", "slepička", "čepička",
              "spicka", "krupicka", "slepicka", "cepicka",
              "špičková", "krupičková", "slepičková", "čepičková",
              "spickova", "krupickova", "slepickova", "cepickova"},
    "picka": {"špička", "krupička", "slepička", "čepička",
              "spicka", "krupicka", "slepicka", "cepicka",
              "špičková", "krupičková", "slepičková", "čepičková",
              "spickova", "krupickova", "slepickova", "cepickova"},
    "pich": {"pospichal", "pospichalova"},
    "dick": {"vodicka", "vodickova", "brdicka", "brdickova"},
    "stet": {"stetina", "stetinova", "brdicka", "brdickova"},
    # "dalsi_slovo": {"povolené_příjmení"},
}

PROFANITY_PROFESSION_WHITELIST = {
    "hovn": {"knihovnice", "knihovnik", "duchovni", "knihovny"},
    "dick": {"ridic", "ridicka", "metodik", "metodicka"},
    "anal": {"analytik", "analyticka", "analyst", "data analyst"},
}

# ============================================================================
# SUSPICIOUS LIST — word boundary matching
# ============================================================================

SUSPICIOUS_LIST = [
    # Komunisté / nacisté / váleční zločinci / diktatori
    "gottwald", "jakes", "eichmann", "goebbels", "himmler", "hitler",
    "Ahmed Abdallah Abderemane",
    "Isaias Afwerki", "Umar al-Bašír",
    "Hejdar Alijev",     "Ilham Alijev",
    "Idi Amin",     "Ion Antonescu",
    "Bašár al-Asad",     "Háfiz al-Asad",
    "Mustafa Kemal Atatürk",     "Hastings Kamuzu Banda",
    "Muhammad Siad Barre",     "Fulgencio Batista",
    "Gurbanguly Berdimuhamedow",     "Serdar Berdimuhamedow",
    "Reynaldo Bignone",     "Zín Abidín bin Alí",
    "Paul Biya",     "Jean-Bédel Bokassa",
    "Simón Bolívar",     "Napoleon Bonaparte",
    "Omar Bongo",     "Habíb Burgiba",
    "Dési Bouterse",     "Marcello Caetano",
    "Moussa Dadis Camara",     "Tiburcio Carías Andino",
    "Óscar Carmona",     "Rafael Carrera y Turcios",
    "Humberto de Alencar Castelo Branco",
    "Fidel Castro",     "Raúl Castro",
    "Louis-Eugène Cavaignac",     "Nicolae Ceaușescu",
    "Jumdžágín Cedenbal",     "Cch’-si",
    "Blaise Compaoré",     "Lansana Conté",
    "Artur da Costa e Silva",     "Oliver Cromwell",
    "Čankajšek",     "Chorlogín Čojbalsan",
    "Čon Du-hwan",     "Idriss Déby",
    "Porfirio Díaz",     "Abdou Diouf",
    "Samuel Doe",     "Engelbert Dollfuss",
    "José Eduardo dos Santos",     "François Duvalier",
    "Jean-Claude Duvalier",     "Ebroin",
    "Enver Paša",     "Recep Tayyip Erdoğan",
    "João Figueiredo",     "Francisco Franco",
    "Alberto Fujimori",     "Leopoldo Galtieri",
    "Giuseppe Garibaldi",     "Maumoon Abdul Gayoom",
    "Juan Vicente Gómez",     "Yakubu Gowon",
    "Hissène Habré",     "Juvénal Habyarimana",
    "Hassanal Bolkiah",     "Adolf Hitler",
    "Ho Či Min",     "Enver Hodža",
    "Erich Honecker",     "Miklós Horthy",
    "Hun Sen",     "Saddám Husajn",
    "Alí Chámeneí",     "Hugo Chávez",
    "Rúholláh Chomejní",     "Nikita Sergejevič Chruščov",
    "Chu Ťin-tchao",     "I Sung-man",
    "Yahya Jammeh",     "Wojciech Jaruzelski",
    "Dawda Jawara",     "János Kádár",
    "Muammar Kaddáfí",     "Ramzan Kadyrov",
    "Paul Kagame",     "Islam Karimov",
    "Kim Čong-il",     "Kim Čong-un",
    "Kim Ir-sen",     "Vladimir Iljič Lenin",
    "Roberto Marcelo Levingston",
    "Li Kuang-jao",
    "Alexandr Lukašenko",
    "Nicolás Maduro",
    "Gerardo Machado",
    "Mao Ce-tung",
    "Ferdinand Marcos",
    "Mengistu Haile Mariam",
    "Emílio Garrastazu Médici",
    "Ioannis Metaxas",
    "Michel Micombero",
    "Slobodan Milošević",
    "Min Aun Hlain",
    "Mobutu Sese Seko",
    "Efraín Ríos Montt",
    "Husní Mubárak",
    "Robert Mugabe",
    "Yoweri Museveni",
    "Benito Mussolini",
    "Parvíz Mušaraf",
    "Napoleon III.",
    "Gamál Násir",
    "Nursultan Nazarbajev",
    "Francisco Macias Nguema",
    "Saparmurat Nijazov",
    "Kwame Nkrumah",
    "Manuel Noriega",
    "Teodoro Obiang Nguema Mbasogo",
    "Nobunaga Oda",
    "Daniel Ortega",
    "José Antonio Páez",
    "Muhammad Rezá Pahlaví",
    "Rezá Šáh Pahlaví",
    "Pak Čong-hui",
    "Georgios Papadopulos",
    "Ante Pavelić",
    "Juan Perón",
    "Pchu I",
    "Pibul Songgram",
    "Józef Piłsudski",
    "Augusto Pinochet",
    "Pol Pot",
    "Miguel Primo de Rivera",
    "Vladimir Putin",
    "Emómalí-ji Rahmón",
    "Maximilien Robespierre",
    "António de Oliveira Salazar",
    "Alí Abdalláh Sálih",
    "Antonio López de Santa Anna",
    "Denis Sassou-Nguesso",
    "Haile Selassie I.",
    "Si Ťin-pching",
    "Abd al-Fattáh as-Sísí",
    "Antanas Smetona",
    "Anastasio Somoza Debayle",
    "Anastasio Somoza García",
    "Somozové",
    "Josif Vissarionovič Stalin",
    "Alfredo Stroessner",
    "Suharto",
    "Sukarno",
    "Teng Siao-pching",
    "Than Šwei",
    "Ťiang Ce-min",
    "Jozef Tiso",
    "Josip Broz Tito",
    "Hideki Tódžó",
    "Hidejoši Tojotomi",
    "Iejasu Tokugawa",
    "Ahmed Sékou Touré",
    "Rafael Trujillo",
    "Walter Ulbricht",
    "Muhammad Umar",
    "Getúlio Vargas",
    "Juan Velasco Alvarado",
    "José María Velasco Ibarra",
    "Jorge Rafael Videla",
    "Roberto Eduardo Viola",
    "Todor Živkov", "caesar",
    "Appius Claudius Caecus",
    "Aulus Postumius Tubertus",
    "Julius Caesar",
    "Cincinnatus",
    "Gaius Duilius",
    "Marcus Furius Camillus",
    "Quintus Fabius Maximus Cunctator",
    "Sulla",
    # Kontroverzní  postavy
    "epstein", "guevara", "klaus", "milos zeman", "heydrich", "himmler", "goring", "hess", 
    "bormann", "frank", "mao", "mengele", "speer", "stalin", "lenin", "breznev", 
     "ribbentrop", "lukasenko", "ceaușescu", "Piłsudski",
    # Žijící politici jako falešné jméno
    "babis", "orban", "trump", "okamura", "rajchl", "macinka", "turek", "fial",
    "klempir", "putin",     
    # Smyšlená jména ze starých dat
    "troubovic", "uzdichcal", "michopulos", "tesiprdel", "lepsozmrdi", "pavek",
    "troub",
    # k vládě
    "petikoal", "petidemoli",
    # Rasistické
    "negr", "neger",
    # Sexuální / urážlivé
    "prostitut", "porno", "raper",
    # Hanlivé role jako falešné jméno
    "diktator", "komunis", "kapitan", "samuraj", "gambler", "alkoholik",
    # Ze starých dat
    "stalingrad", "rozvedcik", "tajemnik",
    "zlodej", "zloděj", "okradeny", "blbec",
    "srackou", "dohajzlu",
    "picin",
    "moloch",
    # z word dokumentu
    "cocain", "kokain" , "herak" ,"heroi", "cannabis", "cern", "pras", "opic", "opil", "alkohol", "alcohol", "drog", "marihuan", "marijuan", "marjan", "pernik", "extaz", "cigar", "zloci", "lepsoli",
    "zrz", "minar", "chvilk", "havl", "kavar", "hnus",
    # Anglické
    "slave", "hater", "monster", "sigma", "sex", 
    # ct related
    "media", "popla", "koncesion", "televiz", "radio", 
    # slechticke tituly
    "kral", "kralovna", "princ", "princezna", "baron",
    # hodnosti
    "major", "general",
]

# ČT: obklopeno mezerami/krajem stringu
CT_PATTERN = re.compile(r"(?<!\S)ct(?!\S)", re.IGNORECASE)

# CRo: obklopeno mezerami/krajem stringu
CRO_PATTERN = re.compile(r"(?<!\S)cro(?!\S)", re.IGNORECASE)

# Mch: obklopeno mezerami/krajem stringu
MCH_PATTERN = re.compile(r"(?<!\S)mch(?!\S)", re.IGNORECASE)

# ============================================================================
# VERB PATTERNS — slovesa nemají co dělat v jméně/městě/povolání
# ============================================================================

VERB_WORDS = [
    " je ", " jsou ", " byl ", " byla ", " bylo ", " bude ",
    " chci ", " chceš ", " chceme ", " chcete ",
    " nechci ", " nechcete ",
    " zruš ", " zrušte ", " zrušit ",
    " dejte ", " dej ",
    " vraťte ", " vrať ",
    " volte ", " nevolte ", " hlasujte ",
    " podpořte ", " podporte ",
    " miluj ", " milujte ", " milovat ",
    " nenávidím ", " nenavidim ",
    " běžte ", " jděte ",
]

VERB_SUFFIXES = ["ujte", "ejte", "ějte"]

# ============================================================================
# PATTERN BUILDING
# ============================================================================

def build_profanity_pattern(word: str):
    w = normalize(word)
    if w in SHORT_PROFANITY or len(w) <= 3:
        return re.compile(r"(?<![a-z])" + re.escape(w) + r"(?![a-z])", re.IGNORECASE)
    return re.compile(re.escape(w), re.IGNORECASE)

def build_suspicious_pattern(word: str):
    w = normalize(word)
    return re.compile(r"(?<!\w)" + re.escape(w) + r"(?!\w)", re.IGNORECASE)

PROFANITY_PATTERNS  = {w: build_profanity_pattern(w)  for w in PROFANITY_LIST}
SUSPICIOUS_PATTERNS = {w: build_suspicious_pattern(w) for w in SUSPICIOUS_LIST}

# ============================================================================
# MATCHING
# ============================================================================

def find_profanity(text: str) -> list:
    if not text:
        return []
    norm = normalize(text)
    return sorted({w for w, p in PROFANITY_PATTERNS.items() if p.search(norm)})

def find_suspicious(text: str) -> list:
    if not text:
        return []
    norm = " " + normalize(text) + " "
    found = {w for w, p in SUSPICIOUS_PATTERNS.items() if p.search(norm)}
    if CT_PATTERN.search(norm):
        found.add("ČT")
    if CRO_PATTERN.search(norm):
        found.add("ČRo")
    if MCH_PATTERN.search(norm):
        found.add("MCh")
    return sorted(found)

def find_verbs(text: str) -> list:
    if not text:
        return []
    norm_padded = " " + normalize(text) + " "
    found = []
    for verb in VERB_WORDS:
        if " " + normalize(verb.strip()) + " " in norm_padded:
            found.append(verb.strip())
    for suffix in VERB_SUFFIXES:
        if re.search(rf"\w+{re.escape(suffix)}\b", norm_padded):
            found.append(f"-{suffix}")
    return sorted(set(found))

def apply_profanity_whitelist(matches: list, jmeno: str, mesto: str, povolani: str) -> list:
    filtered = []
    mesto_norm = normalize(mesto)
    jmeno_norm = normalize(jmeno)
    povolani_norm = normalize(povolani)
    for word in matches:
        word_norm = normalize(word)
        # Whitelist podle města
        if word_norm in PROFANITY_CITY_WHITELIST:
            allowed_cities = {normalize(m) for m in PROFANITY_CITY_WHITELIST[word_norm]}
            if mesto_norm in allowed_cities:
                continue   # OK, přeskočíme
        # Whitelist podle příjmení
        if word_norm in PROFANITY_NAME_WHITELIST:
            allowed_names = {normalize(n) for n in PROFANITY_NAME_WHITELIST[word_norm]}
            if any(allowed in jmeno_norm for allowed in allowed_names):
                continue   # OK, přeskočíme
        # Whitelist podle povolani
        if word_norm in PROFANITY_PROFESSION_WHITELIST:
            allowed_professions = {
                normalize(p)
                for p in PROFANITY_PROFESSION_WHITELIST[word_norm]
            }
            if any(ap in povolani_norm for ap in allowed_professions):
                continue
        filtered.append(word)
    return filtered

def check_length(jmeno: str, mesto: str) -> list:
    """Vrátí seznam problémů s délkou, nebo prázdný seznam."""
    issues = []
    jlen = len(re.sub(r"\s+", " ", jmeno.strip()))
    mlen = len(mesto.strip())
    if jlen > 0 and jlen < MIN_JMENO:
        issues.append(f"jméno krátké ({jlen} znaků, min {MIN_JMENO})")
    if jlen > MAX_JMENO:
        issues.append(f"jméno dlouhé ({jlen} znaků, max {MAX_JMENO})")
    if mlen > MAX_MESTO:
        issues.append(f"město dlouhé ({mlen} znaků, max {MAX_MESTO})")
    return issues

# ============================================================================
# XLSX EXPORT
# ============================================================================

FILL_PROFANITY  = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
FILL_SUSPICIOUS = PatternFill(start_color="FFE5B4", end_color="FFE5B4", fill_type="solid")
FILL_VERB       = PatternFill(start_color="D6F5D6", end_color="D6F5D6", fill_type="solid")
FILL_DUPLICATE  = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
FILL_LENGTH     = PatternFill(start_color="CFE2FF", end_color="CFE2FF", fill_type="solid")
FILL_MULTI      = PatternFill(start_color="E8D5FF", end_color="E8D5FF", fill_type="solid")
FILL_HEADER     = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

HEADERS    = ["DONE?", "Číslo podpisu", "Jméno a Příjmení", "Email", "Co s tím?", "Komentář", "Město", "Povolání"]
COL_WIDTHS = [10,      16,              28,                  30,       14,           55,          22,      28]

def choose_fill(komentar: str) -> PatternFill:
    flags = [
        "Sprosté slovo"   in komentar,
        "Podezřelé slovo" in komentar,
        "Sloveso"         in komentar,
        "Duplicita"       in komentar,
        "Podezřelá délka" in komentar,
    ]
    if sum(flags) > 1:
        return FILL_MULTI
    if flags[0]: return FILL_PROFANITY
    if flags[1]: return FILL_SUSPICIOUS
    if flags[2]: return FILL_VERB
    if flags[4]: return FILL_LENGTH
    return FILL_DUPLICATE

def export_to_xlsx(problems: list, filename: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Problematické podpisy"

    for idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=idx, value=h)
        cell.fill = FILL_HEADER
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for row_idx, p in enumerate(problems, start=2):
        vals = [
            p["done"],
            p["cislo"],
            p["jmeno_prijmeni"],
            p.get("email", ""),
            p.get("co_s_tim", ""),
            p["komentar"],
            p.get("mesto", ""),
            p.get("povolani", ""),
        ]
        fill = choose_fill(p["komentar"])
        for col_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    ws.freeze_panes = "A2"

    # Legenda
    ws2 = wb.create_sheet("Legenda")
    legend = [
        (FILL_PROFANITY,  "Sprosté slovo"),
        (FILL_SUSPICIOUS, "Podezřelé slovo"),
        (FILL_VERB,       "Sloveso v jméně/městě/povolání"),
        (FILL_LENGTH,     f"Podezřelá délka (jméno <{MIN_JMENO} nebo >{MAX_JMENO}, město >{MAX_MESTO})"),
        (FILL_DUPLICATE,  "Duplicita (jméno + město + povolání)"),
        (FILL_MULTI,      "Více kategorií"),
    ]
    ws2.cell(1, 1).value = "Barva"
    ws2.cell(1, 2).value = "Význam"
    for i, (fill, label) in enumerate(legend, start=2):
        ws2.cell(i, 1).fill = fill
        ws2.cell(i, 2).value = label
    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 55

    wb.save(filename)
    print(f"\n💾 Uloženo: {filename}  ({len(problems)} problémů)")

# ============================================================================
# ANALYZE
# ============================================================================

def compute_duplicate_numbers(signatures: list) -> tuple:
    """Vrátí (set duplicitních čísel, slovník skupina→čísla)."""
    dup_map = defaultdict(list)
    for sig in signatures:
        key = (
            normalize(sig["jmeno_prijmeni"]),
            normalize(sig["mesto"]),
            normalize(sig.get("povolani", "")),
        )
        if key[0]:
            dup_map[key].append(sig["cislo"])
    groups = {k: v for k, v in dup_map.items() if len(set(v)) > 1}
    numbers = {n for vals in groups.values() for n in vals}
    return numbers, groups

def analyze(signatures: list) -> list:
    print(f"  Počítám duplicity...")
    dup_numbers, dup_groups = compute_duplicate_numbers(signatures)

    # Obrácená mapa: cislo → klíč skupiny (pro rychlé dohledání sourozenců)
    cislo_to_key = {}
    for key, vals in dup_groups.items():
        for cislo in vals:
            cislo_to_key[cislo] = key

    print(f"  Duplicitních čísel: {len(dup_numbers)}")
    problems = []
    for sig in signatures:
        combined = " ".join([
            sig["jmeno_prijmeni"],
            sig["mesto"],
            sig.get("povolani", ""),
            sig.get("email", ""),
        ])

        reasons, comments = [], []

        prof = find_profanity(combined)
        prof = apply_profanity_whitelist(prof, sig["jmeno_prijmeni"], sig["mesto"], sig["povolani"]) 
        if prof:
            reasons.append("Sprosté slovo")
            comments.append(f"Profanity: {', '.join(prof)}")

        susp = find_suspicious(combined)
        if susp:
            reasons.append("Podezřelé slovo")
            comments.append(f"Suspicious: {', '.join(susp)}")

        verbs = find_verbs(combined)
        if verbs:
            reasons.append("Sloveso")
            comments.append(f"Sloveso: {', '.join(verbs)}")

        length_issues = check_length(sig["jmeno_prijmeni"], sig["mesto"])
        if length_issues:
            reasons.append("Podezřelá délka")
            comments.append(f"Délka: {', '.join(length_issues)}")

        if sig["cislo"] in dup_numbers:
            reasons.append("Duplicita")
            key = cislo_to_key[sig["cislo"]]
            siblings = [n for n in dup_groups[key] if n != sig["cislo"]]
            if siblings:
                comments.append(f"Duplicitní č.: {', '.join(siblings)}")

        if reasons:
            problems.append({
                "done":           "FALSE",
                "cislo":          sig["cislo"],
                "jmeno_prijmeni": sig["jmeno_prijmeni"],
                "email":          sig.get("email", ""),
                "co_s_tim":       "",
                "komentar":       " | ".join(reasons + comments),
                "mesto":          sig["mesto"],
                "povolani":       sig.get("povolani", ""),
            })

    return problems

# ============================================================================
# CSV NAČTENÍ
# ============================================================================

def load_csv(filepath: str) -> list:
    rows = []
    with open(filepath, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows

def load_and_merge_csvs(filepaths: list) -> list:
    """Načte více CSV souborů a deduplikuje podle čísla podpisu."""
    seen = set()
    merged = []
    for fp in filepaths:
        rows = load_csv(fp)
        new = 0
        for row in rows:
            cislo = row.get("cislo", "")
            if cislo and cislo not in seen:
                seen.add(cislo)
                merged.append(row)
                new += 1
        print(f"   {Path(fp).name}: {len(rows)} řádků, přijato {new} nových")
    merged.sort(key=lambda r: int(r["cislo"]) if r["cislo"].isdigit() else 0)
    return merged

def find_all_csvs() -> list:
    """Najde všechny petice_data_*.csv v aktuální složce, seřazené dle názvu."""
    return sorted(str(p) for p in Path(".").glob("petice_data_*.csv"))

# ============================================================================
# MAIN
# ============================================================================

def main():
    print("=" * 60)
    print("  PETICE ANALYZER — ruceprycodmedii.cz")
    print("=" * 60)

    # --- Výběr režimu ---
    print("""
[1] Analyzovat jeden soubor
[2] Sloučit všechny petice_data_*.csv ve složce a analyzovat
[3] Vybrat soubory ručně
""")
    while True:
        mode = input("Zvol možnost [1/2/3]: ").strip()
        if mode in ("1", "2", "3"):
            break
        print("  ❌ Zadej 1, 2 nebo 3.")

    # --- Načtení podle režimu ---
    if mode == "1":
        all_csvs = find_all_csvs()
        if all_csvs:
            print(f"\nNejnovější soubor: {all_csvs[-1]}")
            answer = input("Použít tento soubor? [Y/n]: ").strip().lower()
            csv_file = all_csvs[-1] if answer in ("", "y", "yes") else input("Cesta k souboru: ").strip().strip('"')
        else:
            csv_file = input("\nCesta k souboru: ").strip().strip('"')
        if not Path(csv_file).exists():
            print(f"❌ Soubor nenalezen: {csv_file}")
            return
        print(f"\n📂 Načítám: {csv_file}")
        signatures = load_csv(csv_file)
        print(f"   Načteno {len(signatures)} podpisů")

    elif mode == "2":
        all_csvs = find_all_csvs()
        if not all_csvs:
            print("❌ Žádné petice_data_*.csv soubory nenalezeny.")
            return
        print(f"\n📂 Nalezeno {len(all_csvs)} souborů, slučuji:")
        signatures = load_and_merge_csvs(all_csvs)
        print(f"   Celkem po sloučení: {len(signatures)} unikátních podpisů")

    else:  # mode == "3"
        print("\nZadávej cesty k souborům (prázdný řádek = hotovo):")
        paths = []
        while True:
            p = input(f"  Soubor {len(paths)+1}: ").strip().strip('"')
            if not p:
                break
            if Path(p).exists():
                paths.append(p)
                print(f"    ✅ přidán")
            else:
                print(f"    ❌ soubor nenalezen, zkus znovu")
        if not paths:
            print("❌ Žádné soubory nebyly zadány.")
            return
        print(f"\n📂 Slučuji {len(paths)} souborů:")
        signatures = load_and_merge_csvs(paths)
        print(f"   Celkem po sloučení: {len(signatures)} unikátních podpisů")

    # --- Analýza ---
    print("\n🔎 Analyzuji...")
    problems = analyze(signatures)
    print(f"   Nalezeno problémů: {len(problems)}")

    cats = ["Sprosté slovo", "Podezřelé slovo", "Sloveso", "Podezřelá délka", "Duplicita"]
    for cat in cats:
        n = sum(1 for p in problems if cat in p["komentar"])
        print(f"   — {cat}: {n}")

    export_to_xlsx(problems, OUTPUT_XLSX)


if __name__ == "__main__":
    main()

